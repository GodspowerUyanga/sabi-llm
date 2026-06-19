"""
Local tools for the Sabi agent.

This is the load-bearing cross-disciplinary pairing for the Corporate /
Enterprise track: knowledge work (language) x quantitative reasoning (exact
maths over business data). Small LLMs are unreliable at arithmetic; routing
numbers through a deterministic tool is what makes Sabi trustworthy for SME
finance and operations — and it directly protects the Accuracy score.

All tools are pure-Python, offline, and side-effect free.

Tools:
- calc      : evaluate an arithmetic / financial expression safely (no eval()).
- aggregate : sum / mean / min / max / count over a column of a CSV in the corpus.
- search_docs: retrieve from the local RAG index (wired in by the agent).
"""
from __future__ import annotations

import ast
import csv
import json
import math
import operator
from pathlib import Path
from typing import Any, Callable

# ----------------------------------------------------------------------------
# Safe arithmetic evaluator (AST-based; never uses eval/exec)
# ----------------------------------------------------------------------------
_BIN_OPS: dict[type, Callable] = {
    ast.Add: operator.add, ast.Sub: operator.sub, ast.Mult: operator.mul,
    ast.Div: operator.truediv, ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod, ast.Pow: operator.pow,
}
_UNARY_OPS: dict[type, Callable] = {ast.UAdd: operator.pos, ast.USub: operator.neg}
_FUNCS: dict[str, Callable] = {
    "abs": abs, "round": round, "min": min, "max": max, "sum": sum,
    "sqrt": math.sqrt, "log": math.log, "exp": math.exp, "floor": math.floor,
    "ceil": math.ceil, "pow": pow,
}
_CONSTS = {"pi": math.pi, "e": math.e}


def safe_eval(expr: str) -> float:
    """Evaluate a math expression with a strict allow-list of node types."""
    tree = ast.parse(expr, mode="eval")

    def _ev(node: ast.AST) -> Any:
        if isinstance(node, ast.Expression):
            return _ev(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return node.value
            raise ValueError("only numeric constants allowed")
        if isinstance(node, ast.BinOp) and type(node.op) in _BIN_OPS:
            return _BIN_OPS[type(node.op)](_ev(node.left), _ev(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _UNARY_OPS:
            return _UNARY_OPS[type(node.op)](_ev(node.operand))
        if isinstance(node, ast.Name) and node.id in _CONSTS:
            return _CONSTS[node.id]
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id in _FUNCS:
            args = [_ev(a) for a in node.args]
            return _FUNCS[node.func.id](*args)
        if isinstance(node, (ast.List, ast.Tuple)):
            return [_ev(e) for e in node.elts]
        raise ValueError(f"unsupported expression element: {type(node).__name__}")

    return _ev(tree)


# ----------------------------------------------------------------------------
# Tool implementations
# ----------------------------------------------------------------------------
def tool_calc(expression: str, **_: Any) -> dict[str, Any]:
    try:
        value = safe_eval(str(expression))
        return {"expression": expression, "result": value}
    except Exception as exc:
        return {"error": f"could not evaluate '{expression}': {exc}"}


import re as _re

# A run of characters that could form an arithmetic expression.
_ARITH_RUN = _re.compile(r"[0-9.,()+\-*/%₦$ ]+")


def extract_expression(text: str) -> str | None:
    """If *text* contains a clear arithmetic expression, return a clean,
    evaluatable version of it; else None.

    Used by the agent to guarantee exact arithmetic: explicit calculation
    requests are computed by the calc tool rather than left to the model.
    Guards against false positives like number ranges ("5-10 days").
    """
    if not text:
        return None
    best: str | None = None
    for m in _ARITH_RUN.finditer(text):
        cleaned = (m.group(0).replace(",", "").replace("₦", "").replace("$", "")
                   .replace("%", "").replace(" ", "").strip().strip("+-*/"))
        if not cleaned or not any(op in cleaned for op in "+-*/"):
            continue
        if not any(c.isdigit() for c in cleaned):
            continue
        if cleaned.count("(") != cleaned.count(")"):
            continue
        # Only treat as a real calculation if there's a strong signal — avoids
        # firing on ranges/IDs like "5-10" or "172-16".
        nums = _re.findall(r"\d+", cleaned)
        strong = ("*" in cleaned or "/" in cleaned or "(" in cleaned or "." in cleaned
                  or any(len(n) >= 4 for n in nums)
                  or sum(cleaned.count(op) for op in "+-*/") >= 2)
        if not strong:
            continue
        try:
            safe_eval(cleaned)
        except Exception:
            continue
        if best is None or len(cleaned) > len(best):
            best = cleaned
    return best


def _to_number(s) -> float | None:
    s = str(s).strip().replace(",", "").replace("$", "").replace("₦", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def _read_table(path: Path) -> tuple[list[str], list[dict]]:
    """Read a CSV or Excel file into (headers, list-of-row-dicts)."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(next(rows_iter, []))]
        rows = []
        for r in rows_iter:
            rows.append({headers[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(headers)})
        wb.close()
        return headers, rows
    # CSV / TSV
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        reader = csv.DictReader(fh)
        headers = list(reader.fieldnames or [])
        rows = list(reader)
    return headers, rows


def tool_aggregate(file: str, column: str, op: str = "sum", corpus_dir: str | Path = ".", **_: Any) -> dict[str, Any]:
    """Aggregate a numeric column from a CSV/Excel file in the corpus directory."""
    op = op.lower()
    if op not in {"sum", "mean", "avg", "average", "min", "max", "count"}:
        return {"error": f"unsupported op '{op}'"}
    path = Path(corpus_dir) / file
    if not path.exists():
        matches = list(Path(corpus_dir).rglob(f"{file}*"))
        if not matches:
            return {"error": f"file '{file}' not found in corpus"}
        path = matches[0]
    try:
        headers, rows = _read_table(path)
    except Exception as exc:
        return {"error": f"failed reading {path.name}: {exc}"}
    if column not in headers:
        return {"error": f"column '{column}' not in {path.name}; columns: {headers}"}
    values = [n for row in rows if (n := _to_number(row.get(column, ""))) is not None]
    if not values:
        return {"error": f"no numeric values in column '{column}'"}
    result = {
        "sum": sum(values), "mean": sum(values) / len(values),
        "avg": sum(values) / len(values), "average": sum(values) / len(values),
        "min": min(values), "max": max(values), "count": len(values),
    }[op]
    return {"file": path.name, "column": column, "op": op, "result": round(result, 4), "n": len(values)}


# ---------------------------------------------------------------------------
# Deterministic data analysis (used by the agent so the model never does maths
# over a table itself — this prevents hallucinated totals on small models).
# ---------------------------------------------------------------------------
_AGG_INTENT = [
    (("total", "sum", "altogether", "combined", "overall"), "sum"),
    (("average", "mean", "avg"), "mean"),
    (("how many", "count", "number of"), "count"),
    (("highest", "max", "maximum", "largest", "most", "top", "biggest"), "max"),
    (("lowest", "min", "minimum", "smallest", "least"), "min"),
]


def _detect_op(text: str) -> str | None:
    t = text.lower()
    for keys, op in _AGG_INTENT:
        if any(k in t for k in keys):
            return op
    return None


def _numeric_columns(headers: list[str], rows: list[dict]) -> list[str]:
    cols = []
    for h in headers:
        vals = [_to_number(r.get(h, "")) for r in rows[:20]]
        if vals and sum(1 for v in vals if v is not None) >= max(1, len(vals) // 2):
            cols.append(h)
    return cols


def _match_column(text: str, numeric_cols: list[str]) -> str | None:
    t = text.lower()
    for col in numeric_cols:
        tokens = [tok for tok in col.lower().replace("-", "_").split("_") if len(tok) >= 3]
        for tok in tokens:
            if tok in t or tok.rstrip("s") in t:
                return col
    return None


def auto_aggregate(text: str, corpus_dir: str | Path):
    """If *text* is a clear data question (total/average/etc. of a column in a
    spreadsheet), compute it deterministically. If the user names a single group
    (e.g. "total revenue in Port Harcourt"), the result is filtered to that group;
    otherwise a per-group breakdown is returned.
    """
    op = _detect_op(text)
    if not op:
        return None
    for _, path, headers, rows in _rank_tables(text, corpus_dir):
        numeric_cols = _numeric_columns(headers, rows)
        value_col = _match_column(text, numeric_cols)
        if not value_col:
            continue
        group_col = next((h for h in headers if h not in numeric_cols), None)
        flt = _detect_filter(text, headers, rows, numeric_cols)

        def _agg_over(rs):
            vals = [n for r in rs if (n := _to_number(r.get(value_col, ""))) is not None]
            if not vals:
                return None, 0
            res = {"sum": sum(vals), "mean": sum(vals) / len(vals), "count": len(vals),
                   "max": max(vals), "min": min(vals)}[op]
            return res, len(vals)

        # Single-group filter ("... for Port Harcourt")
        if flt:
            col, val = flt
            sub = [r for r in rows if str(r.get(col, "")).strip() == val]
            res, n = _agg_over(sub)
            if res is not None:
                return {"op": op, "file": path.name, "value_column": value_col,
                        "group_column": group_col, "result": round(res, 4),
                        "breakdown": [], "n": n, "filter_column": col, "filter_value": val}

        res, n = _agg_over(rows)
        if res is None:
            continue
        breakdown = []
        if group_col:
            groups: dict[str, list[float]] = {}
            for r in rows:
                v = _to_number(r.get(value_col, ""))
                if v is not None:
                    groups.setdefault(str(r.get(group_col, "")), []).append(v)
            for g, vs in groups.items():
                gv = {"sum": sum(vs), "mean": sum(vs) / len(vs), "count": len(vs),
                      "max": max(vs), "min": min(vs)}[op]
                breakdown.append((g, gv))
        return {"op": op, "file": path.name, "value_column": value_col,
                "group_column": group_col, "result": round(res, 4),
                "breakdown": breakdown, "n": n, "filter_value": None}
    return None


# ----------------------------------------------------------------------------
# Tool registry + JSON spec rendered into the system prompt
# ----------------------------------------------------------------------------
TOOL_SPECS = [
    {
        "name": "calc",
        "description": "Evaluate an arithmetic expression and return the exact number. Use for any maths.",
        "arguments": {"expression": "string, e.g. '0.075 * 1250000' or '(45000-38000)/38000*100'"},
    },
    {
        "name": "aggregate",
        "description": "Compute sum/mean/min/max/count over a numeric column of a CSV in the company documents.",
        "arguments": {"file": "csv filename", "column": "column name", "op": "sum|mean|min|max|count"},
    },
    {
        "name": "search_docs",
        "description": "Search the company's documents for relevant passages.",
        "arguments": {"query": "string"},
    },
]


def render_tool_specs() -> str:
    lines = []
    for spec in TOOL_SPECS:
        lines.append(
            f"- {spec['name']}: {spec['description']} "
            f"arguments={json.dumps(spec['arguments'])}"
        )
    return "\n".join(lines)


def build_registry(corpus_dir: str | Path, search_docs: Callable[[str], str] | None) -> dict[str, Callable[..., dict]]:
    """Bind tool implementations to runtime context (corpus dir, RAG search)."""
    registry: dict[str, Callable[..., dict]] = {
        "calc": tool_calc,
        "aggregate": lambda **kw: tool_aggregate(corpus_dir=corpus_dir, **kw),
    }
    if search_docs is not None:
        registry["search_docs"] = lambda query, **_: {"query": query, "result": search_docs(query)}
    return registry


def export_to_xlsx(content: str, title: str, out_dir: str | Path) -> str:
    """Write *content* (a summary / notes) into a real .xlsx file and return its name.

    Lines of the form "key: value" become two columns; other lines go in one
    column. Used by the agent when the user asks Sabi to create an Excel sheet.
    """
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = (title[:28] or "Summary")
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for raw in content.split("\n"):
        line = _re.sub(r"^[\-\*•\d\.\)\s]+", "", raw.strip())
        line = line.replace("**", "").strip()
        if not line:
            continue
        if ":" in line and len(line.split(":", 1)[0]) <= 45:
            k, v = line.split(":", 1)
            ws.append([k.strip(), v.strip()])
        else:
            ws.append([line])
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 65
    slug = (_re.sub(r"[^A-Za-z0-9]+", "_", title)[:30].strip("_") or "summary")
    name = f"{slug}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(out / name))
    return name


# ---------------------------------------------------------------------------
# Pivot tables and spreadsheet analysis (deterministic — never the LLM's maths)
# ---------------------------------------------------------------------------
def _agg(values: list[float], op: str) -> float:
    if not values:
        return 0.0
    if op in ("mean", "avg", "average"):
        return sum(values) / len(values)
    if op == "count":
        return float(len(values))
    if op == "min":
        return min(values)
    if op == "max":
        return max(values)
    return sum(values)  # default sum


def _resolve_table(file: str, corpus_dir) -> Path | None:
    p = Path(corpus_dir) / file
    if p.exists():
        return p
    matches = [m for m in Path(corpus_dir).rglob(f"*{file}*")
               if m.suffix.lower() in (".csv", ".xlsx", ".xls")]
    return matches[0] if matches else None


def pivot_table(file: str, index: str, value: str, agg: str = "sum",
                columns: str | None = None, corpus_dir="."):
    """Build a pivot table: group `value` by `index` (rows) and optional
    `columns`, aggregated with `agg`. Returns a structured dict."""
    path = _resolve_table(file, corpus_dir)
    if not path:
        return {"error": f"file '{file}' not found"}
    headers, rows = _read_table(path)
    for c in [index, value] + ([columns] if columns else []):
        if c not in headers:
            return {"error": f"column '{c}' not in {path.name}; columns: {headers}"}

    idx_values = sorted({str(r.get(index, "")) for r in rows if str(r.get(index, "")) != ""})
    if columns:
        col_values = sorted({str(r.get(columns, "")) for r in rows if str(r.get(columns, "")) != ""})
        grid = {iv: {cv: [] for cv in col_values} for iv in idx_values}
        for r in rows:
            v = _to_number(r.get(value, ""))
            if v is None:
                continue
            iv, cv = str(r.get(index, "")), str(r.get(columns, ""))
            if iv in grid and cv in grid[iv]:
                grid[iv][cv].append(v)
        header = [index] + col_values + (["Total"] if agg == "sum" else [])
        out_rows = []
        for iv in idx_values:
            row = [iv]
            row_all = []
            for cv in col_values:
                vals = grid[iv][cv]
                row.append(_agg(vals, agg)); row_all += vals
            if agg == "sum":
                row.append(sum(row_all))
            out_rows.append(row)
    else:
        groups: dict[str, list[float]] = {}
        for r in rows:
            v = _to_number(r.get(value, ""))
            if v is not None:
                groups.setdefault(str(r.get(index, "")), []).append(v)
        header = [index, f"{agg} of {value}"]
        out_rows = [[iv, _agg(groups.get(iv, []), agg)] for iv in idx_values]

    return {"file": path.name, "index": index, "value": value, "agg": agg,
            "columns": columns, "header": header, "rows": out_rows}


def render_pivot_markdown(piv: dict) -> str:
    if "error" in piv:
        return f"_Could not build the pivot table: {piv['error']}_"
    val = piv["value"]
    title = (f"**Pivot — {piv['agg']} of {val.replace('_',' ')} by "
             f"{piv['index'].replace('_',' ')}"
             + (f" × {piv['columns'].replace('_',' ')}" if piv['columns'] else "")
             + f"** _(from {piv['file']})_")
    head = "| " + " | ".join(str(h).replace("_", " ").title() for h in piv["header"]) + " |"
    sep = "|" + "|".join(["---"] + ["---:"] * (len(piv["header"]) - 1)) + "|"
    lines = [title, "", head, sep]
    grand = []
    for row in piv["rows"]:
        cells = [str(row[0])] + [_fmt_money(val, c) for c in row[1:]]
        lines.append("| " + " | ".join(cells) + " |")
        grand.append(row[1:])
    if piv["agg"] == "sum" and grand:
        totals = [sum(col) for col in zip(*grand)]
        lines.append("| **Total** | " + " | ".join(f"**{_fmt_money(val, t)}**" for t in totals) + " |")
    return "\n".join(lines)


_MONEY = ("ngn", "revenue", "naira", "price", "cost", "amount", "sales", "income", "profit")


def _fmt_money(column: str, value) -> str:
    try:
        value = float(value)
    except (TypeError, ValueError):
        return str(value)
    body = f"{value:,.0f}" if float(value).is_integer() else f"{value:,.2f}"
    return f"₦{body}" if any(h in column.lower() for h in _MONEY) else body


def auto_pivot(text: str, corpus_dir):
    """Detect an explicit pivot request and build it. Only triggers on the word
    'pivot', and only over a table whose columns match the dimensions named."""
    t = text.lower()
    if "pivot" not in t:
        return None
    for path in _pick_table(text, corpus_dir):
        if path.suffix.lower() not in (".csv", ".xlsx", ".xls"):
            continue
        try:
            headers, rows = _read_table(path)
        except Exception:
            continue
        if not rows:
            continue
        numeric = _numeric_columns(headers, rows)
        value = _match_column(text, numeric) or (numeric[0] if numeric else None)
        text_cols = [h for h in headers if h not in numeric]
        mentioned = [c for c in text_cols
                     if any(tok in t for tok in c.lower().replace("-", "_").split("_") if len(tok) >= 3)]
        if not mentioned:        # the named dimensions aren't in this file — try the next
            continue
        index = mentioned[0]
        columns = mentioned[1] if len(mentioned) >= 2 else None
        op = _detect_op(text) or "sum"
        if value and index:
            piv = pivot_table(path.name, index, value, op, columns, corpus_dir)
            if "error" not in piv:
                return piv
    return None


def analyze_table(file: str, corpus_dir="."):
    """Summary statistics for every numeric column in a spreadsheet."""
    path = _resolve_table(file, corpus_dir)
    if not path:
        return {"error": f"file '{file}' not found"}
    headers, rows = _read_table(path)
    numeric = _numeric_columns(headers, rows)
    stats = []
    for col in numeric:
        vals = [n for r in rows if (n := _to_number(r.get(col, ""))) is not None]
        if vals:
            stats.append({"column": col, "count": len(vals), "sum": sum(vals),
                          "mean": sum(vals) / len(vals), "min": min(vals), "max": max(vals)})
    return {"file": path.name, "rows": len(rows), "columns": headers, "stats": stats}


# ---------------------------------------------------------------------------
# Deterministic table Q&A — filter / list / count / who-owes / highest-lowest.
# This is what makes Sabi accurate on spreadsheets: the code reads the actual
# cells and computes the answer. The language model NEVER does this arithmetic.
# ---------------------------------------------------------------------------
_OWE_WORDS = ("owe", "owing", "owed", "debt", "debtor", "outstanding", "unpaid",
              "balance", "due", "arrears", "pending", "owning")
_PAID_WORDS = ("paid", "cleared", "settled", "complete", "full", "completed")


def _find_col(headers, hints, avoid=()):
    for h in headers:
        hl = h.lower()
        if any(a in hl for a in avoid):
            continue
        if any(hh in hl for hh in hints):
            return h
    return None


def _cols(headers):
    name = (_find_col(headers, ("name",), avoid=("id",))
            or _find_col(headers, ("customer", "client", "debtor"), avoid=("id",)))
    paid = _find_col(headers, ("paid", "payment", "deposit"))
    balance = _find_col(headers, ("balance", "outstanding", "owed", "arrears", "due"))
    amount = _find_col(headers, ("sale amount", "amount", "total", "price", "value", "invoice amount"),
                       avoid=("paid", "balance"))
    status = _find_col(headers, ("status", "state"))
    product = _find_col(headers, ("product", "item", "description", "goods", "device"))
    return {"name": name, "paid": paid, "balance": balance, "amount": amount,
            "status": status, "product": product}


def _is_owing(r, c):
    if c["status"]:
        s = str(r.get(c["status"], "")).lower()
        if any(w in s for w in ("own", "owe", "unpaid", "pending", "outstanding", "partial", "debt")):
            return True
        if any(w in s for w in _PAID_WORDS):
            if c["balance"]:
                b = _to_number(r.get(c["balance"]))
                return b is not None and b > 0
            return False
    if c["balance"]:
        b = _to_number(r.get(c["balance"]))
        if b is not None:
            return b > 0
    if c["paid"] and c["amount"]:
        p, a = _to_number(r.get(c["paid"])), _to_number(r.get(c["amount"]))
        if p is not None and a is not None:
            return p < a
    return False


def _owe_amount(r, c):
    if c["balance"]:
        b = _to_number(r.get(c["balance"]))
        if b is not None:
            return b
    if c["paid"] and c["amount"]:
        p, a = _to_number(r.get(c["paid"])), _to_number(r.get(c["amount"]))
        if p is not None and a is not None:
            return a - p
    return _to_number(r.get(c["amount"])) if c["amount"] else None


def _table_files(corpus_dir):
    out = []
    for p in sorted(Path(corpus_dir).rglob("*")):
        if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xls"):
            out.append(p)
    return out


def _pick_table(text, corpus_dir):
    files = _table_files(corpus_dir)
    t = text.lower()
    files.sort(key=lambda p: 0 if any(tok in t for tok in p.stem.lower().replace("-", "_").split("_")
                                      if len(tok) >= 3) else 1)
    return files


def _debtor_rows(rows, c):
    out = []
    for r in rows:
        if _is_owing(r, c):
            out.append(r)
    return out


def _list_table(rows, c, value_label="Balance"):
    header = [c["name"] or "Name"]
    if c["product"]:
        header.append(c["product"])
    header.append(value_label)
    out_rows = []
    for r in rows:
        row = [str(r.get(c["name"], "")).strip() or "—"]
        if c["product"]:
            row.append(str(r.get(c["product"], "")).strip())
        amt = _owe_amount(r, c)
        row.append(amt if amt is not None else "")
        out_rows.append(row)
    return header, out_rows


def _render_rows(header, rows, money_cols=()):
    head = "| " + " | ".join(str(h).replace("_", " ").title() for h in header) + " |"
    sep = "|" + "|".join(["---"] + ["---:"] * (len(header) - 1)) + "|"
    lines = [head, sep]
    for r in rows:
        cells = []
        for i, v in enumerate(r):
            if i in money_cols or (i > 0 and isinstance(v, (int, float))):
                cells.append(_fmt_money(str(header[i]), v))
            else:
                cells.append(str(v))
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def query_table(text: str, corpus_dir) -> dict | None:
    """Answer a structured question about a spreadsheet, deterministically.

    Handles: who is owing / list debtors, how many are owing, total owed,
    highest / lowest payment or amount. Returns None if it doesn't apply.
    """
    t = text.lower()
    owe = any(w in t for w in _OWE_WORDS)
    wants_count = any(w in t for w in ("how many", "number of", "count", "how much in total are"))
    wants_max = any(w in t for w in ("highest", "largest", "most", "maximum", "top", "biggest", "max"))
    wants_min = any(w in t for w in ("lowest", "smallest", "least", "minimum", "min"))
    wants_total = any(w in t for w in ("total", "sum", "altogether", "combined")) and owe
    is_payment = any(w in t for w in ("paid", "payment", "pay"))

    for path in _pick_table(text, corpus_dir):
        try:
            headers, rows = _read_table(path)
        except Exception:
            continue
        if not rows:
            continue
        c = _cols(headers)
        numeric = _numeric_columns(headers, rows)
        has_owe_cols = bool(c["name"]) and bool(c["status"] or c["balance"] or (c["paid"] and c["amount"]))

        # Highest / lowest payment or amount
        if (wants_max or wants_min) and not owe:
            if is_payment and not c["paid"]:
                continue  # asked about payments but this table has no payment column
            metric = (c["paid"] if (is_payment and c["paid"]) else _match_column(text, numeric))
            if not metric:
                continue  # don't guess a metric the user didn't name
            best, bestval = None, None
            for r in rows:
                v = _to_number(r.get(metric))
                if v is None:
                    continue
                if bestval is None or (wants_max and v > bestval) or (wants_min and v < bestval):
                    best, bestval = r, v
            if best is None:
                continue
            who = str(best.get(c["name"], "")).strip() if c["name"] else ""
            extra = f" ({best.get(c['product'])})" if c["product"] and best.get(c["product"]) else ""
            label = "highest" if wants_max else "lowest"
            answer = (f"The {label} {metric.lower()} is **{_fmt_money(metric, bestval)}**"
                      + (f", by **{who}**{extra}." if who else "."))
            return {"kind": "value", "markdown": answer, "summary": f"{label} {metric}",
                    "file": path.name}

        if not owe:
            continue
        if not has_owe_cols:
            continue  # this spreadsheet can't answer an owing question — try the next
        debtors = _debtor_rows(rows, c)

        if wants_total:
            total = sum((_owe_amount(r, c) or 0) for r in debtors)
            mc = c["balance"] or c["amount"] or "amount"
            answer = (f"The total outstanding across **{len(debtors)}** "
                      f"customer{'s' if len(debtors) != 1 else ''} is "
                      f"**{_fmt_money(mc, total)}**.")
            return {"kind": "value", "markdown": answer, "summary": "total owed", "file": path.name}

        if wants_count:
            answer = (f"**{len(debtors)}** customer{'s are' if len(debtors) != 1 else ' is'} "
                      f"currently owing.")
            if debtors:
                header, drows = _list_table(debtors, c)
                answer += "\n\n" + _render_rows(header, drows)
            return {"kind": "count", "markdown": answer, "summary": f"{len(debtors)} owing",
                    "file": path.name, "count": len(debtors)}

        # default: list everyone owing (who is owing / list debtors / tabular)
        if not debtors:
            return {"kind": "list", "markdown": "No customers are currently owing — all balances are settled.",
                    "summary": "0 owing", "file": path.name, "header": [], "rows": []}
        header, drows = _list_table(debtors, c)
        intro = (f"**{len(debtors)}** customer{'s are' if len(debtors) != 1 else ' is'} currently owing:\n\n")
        return {"kind": "list", "markdown": intro + _render_rows(header, drows),
                "summary": f"{len(debtors)} owing", "file": path.name,
                "header": header, "rows": drows}
    return None


_SHEET_NAME_RE = _re.compile(
    r"(?:call(?:ed)?|name[d]?|titled?|label(?:ed)?)\s+(?:it\s+|the\s+sheet\s+|this\s+)?[\"']?([A-Za-z][\w-]{0,28})",
    _re.I)


def extract_sheet_name(text: str) -> str | None:
    m = _SHEET_NAME_RE.search(text)
    if m:
        return m.group(1).strip().strip("\"'").title()
    for kw in ("debtor", "debtors"):
        if kw in text.lower():
            return "Debtors"
    return None


def export_table_to_xlsx(header, rows, title, out_dir) -> str:
    """Write a real data table (header + rows) into an .xlsx and return its name."""
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = (title[:28] or "Sheet1")
    ws.append([str(h).replace("_", " ").title() for h in header])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="155E4B")
    for r in rows:
        ws.append([(round(v, 2) if isinstance(v, float) else v) for v in r])
    for i, h in enumerate(header, 1):
        width = max(len(str(h)) + 4, max((len(str(r[i - 1])) for r in rows), default=10) + 3)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(40, width)
    slug = (_re.sub(r"[^A-Za-z0-9]+", "_", title)[:24].strip("_") or "sheet")
    name = f"{slug}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(out / name))
    return name


# ---------------------------------------------------------------------------
# Table viewing + filtering — render the ACTUAL spreadsheet rows deterministically
# so the language model never has to (it truncates rows and drops digits).
# ---------------------------------------------------------------------------
def _fuzzy_contains(value, text) -> float:
    """How strongly `value` appears in `text` (handles typos like Porthacourt)."""
    import difflib
    nv = _re.sub(r"[^a-z0-9]", "", str(value).lower())
    nt = _re.sub(r"[^a-z0-9]", "", str(text).lower())
    if not nv or not nt:
        return 0.0
    if nv in nt:
        return 1.0
    if len(nv) < 3:
        return 0.0
    L = len(nv)
    best = 0.0
    for i in range(0, max(1, len(nt) - L + 1)):
        r = difflib.SequenceMatcher(None, nv, nt[i:i + L]).ratio()
        if r > best:
            best = r
    return best


def _rank_tables(text: str, corpus_dir):
    """Rank spreadsheet files by relevance to the question (filename, columns,
    and mentioned cell values like a region or product name)."""
    t = text.lower()
    out = []
    for p in _table_files(corpus_dir):
        try:
            headers, rows = _read_table(p)
        except Exception:
            continue
        if not rows:
            continue
        score = 0
        for tok in _re.split(r"[_\-. ]+", p.stem.lower()):
            if len(tok) >= 3 and tok in t:
                score += 2
        numeric = _numeric_columns(headers, rows)
        for h in headers:
            for tok in _re.split(r"[_\-. ]+", h.lower()):
                if len(tok) >= 3 and tok in t:
                    score += 1
        for col in headers:
            if col in numeric:
                continue
            uniq = {str(r.get(col, "")).strip() for r in rows}
            if len(uniq) <= 60:
                if any(v and _fuzzy_contains(v, text) >= 0.85 for v in uniq):
                    score += 2
        out.append((score, p, headers, rows))
    out.sort(key=lambda x: -x[0])
    return out


def _detect_filter(text: str, headers, rows, numeric):
    """Find a (column, value) the user is filtering on, e.g. region = Port Harcourt."""
    best, best_ratio, best_len = None, 0.0, 0
    for col in headers:
        if col in numeric:
            continue
        uniq = {str(r.get(col, "")).strip() for r in rows if str(r.get(col, "")).strip()}
        if len(uniq) > 80:
            continue
        for v in uniq:
            if len(_re.sub(r"[^a-z0-9]", "", v.lower())) < 3:
                continue
            ratio = _fuzzy_contains(v, text)
            if ratio >= 0.85 and (ratio > best_ratio or (ratio == best_ratio and len(v) > best_len)):
                best, best_ratio, best_len = (col, v), ratio, len(v)
    return best


_SHOW_RE = _re.compile(
    r"\b(show|list|display|give|see|view|break\s?down|content|contents|full table|"
    r"all the|everything|the data|the table|the rows|tabular|table of|let me see|write (it )?complete|write complete)\b",
    _re.I)
_SHOW_BLOCK_RE = _re.compile(
    r"\b(summar|explain|pivot|owe|owing|debtor|how many|how much|average|mean|"
    r"\btotal\b|sum of|create|export|letter|draft)\b", _re.I)


def show_table(text: str, corpus_dir):
    """Render the real rows of the most relevant spreadsheet, optionally filtered
    to a value the user named (region, product, customer …)."""
    if not _SHOW_RE.search(text) or _SHOW_BLOCK_RE.search(text):
        return None
    ranked = _rank_tables(text, corpus_dir)
    if not ranked or ranked[0][0] == 0:
        return None
    _, path, headers, rows = ranked[0]
    numeric = _numeric_columns(headers, rows)
    flt = _detect_filter(text, headers, rows, numeric)
    shown, note = rows, ""
    if flt:
        col, val = flt
        sub = [r for r in rows if str(r.get(col, "")).strip() == val]
        if sub:
            shown, note = sub, f" — {col.replace('_', ' ')} = **{val}**"
    capped = shown[:100]
    head = "| " + " | ".join(h.replace("_", " ").title() for h in headers) + " |"
    sep = "|" + "|".join(["---"] + ["---:" if h in numeric else "---" for h in headers[1:]]) + "|"
    lines = [f"**{path.name}**{note} — {len(shown)} row(s):", "", head, sep]
    for r in capped:
        lines.append("| " + " | ".join(
            (_fmt_money(h, r.get(h, "")) if h in numeric else str(r.get(h, ""))) for h in headers) + " |")
    if len(capped) > 1 and numeric:
        tot = ["**Total**"]
        for h in headers[1:]:
            if h in numeric:
                s = sum(n for r in capped if (n := _to_number(r.get(h))) is not None)
                tot.append(f"**{_fmt_money(h, s)}**")
            else:
                tot.append("")
        lines.append("| " + " | ".join(tot) + " |")
    if len(shown) > 100:
        lines.append(f"\n_(showing first 100 of {len(shown)} rows)_")
    return {"markdown": "\n".join(lines), "file": path.name, "rows": len(shown),
            "summary": f"{len(shown)} rows from {path.name}"}
