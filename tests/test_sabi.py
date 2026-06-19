"""End-to-end tests for Sabi using the mock backends (no model download needed)."""
import sys
from pathlib import Path

import numpy as np
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from sabi.config import load_config
from sabi.tools import safe_eval, tool_calc, tool_aggregate, build_registry, render_tool_specs
from sabi.embeddings import MockEmbedder, cosine_sim
from sabi.rag import RagIndex, chunk_text
from sabi.languages import detect_language, language_directive
from sabi.memory import MemoryReport, measure_peak
from sabi.model import MockChat
from sabi.agent import SabiAgent


# ---------------------------------------------------------------- config
def test_config_loads_defaults():
    cfg = load_config()
    assert cfg.model.name == "Sabi-1"
    assert cfg.model.ram_budget_gb == 7.0
    assert cfg.model.n_threads >= 1  # auto-resolved


# ---------------------------------------------------------------- tools
@pytest.mark.parametrize("expr,expected", [
    ("0.15 * 52000000", 7800000.0),
    ("(45000-38000)/38000*100", pytest.approx(18.42, abs=0.01)),
    ("sqrt(144) + 2**3", 20.0),
    ("max(3, 9, 5)", 9),
])
def test_safe_eval(expr, expected):
    assert safe_eval(expr) == expected


def test_safe_eval_blocks_code():
    with pytest.raises(Exception):
        safe_eval("__import__('os').system('echo hi')")


def test_tool_calc():
    assert tool_calc(expression="2+2")["result"] == 4
    assert "error" in tool_calc(expression="open('x')")


def test_tool_aggregate_csv():
    corpus = ROOT / "data" / "corpus"
    res = tool_aggregate(file="q1_sales_report.csv", column="revenue_ngn", op="sum", corpus_dir=corpus)
    assert res["result"] == pytest.approx(36690000.0)
    assert res["n"] == 12
    res2 = tool_aggregate(file="q1_sales_report.csv", column="units_sold", op="max", corpus_dir=corpus)
    assert res2["result"] == 1402


def test_render_tool_specs():
    specs = render_tool_specs()
    assert "calc" in specs and "aggregate" in specs and "search_docs" in specs


# ----------------------------------------------------------- embeddings/rag
def test_cosine_sim_identity():
    v = np.array([1.0, 2.0, 3.0])
    mat = np.vstack([v, -v, np.zeros(3)])
    sims = cosine_sim(v, mat)
    assert sims[0] == pytest.approx(1.0, abs=1e-5)
    assert sims[1] == pytest.approx(-1.0, abs=1e-5)


def test_chunking_overlap():
    text = "word " * 400
    chunks = chunk_text(text, size=300, overlap=50)
    assert len(chunks) > 1
    assert all(len(c) <= 320 for c in chunks)


def test_rag_build_and_search(tmp_path):
    embedder = MockEmbedder()
    idx = RagIndex(tmp_path / "idx")
    n = idx.build(ROOT / "data" / "corpus", embedder, chunk_size=400, overlap=80)
    assert n > 0
    qv = embedder.embed(["what is the refund and returns policy"])[0]
    results = idx.search(qv, top_k=3, min_score=0.0)
    assert results
    # persistence round-trip
    idx2 = RagIndex(tmp_path / "idx")
    assert idx2.load()
    assert idx2.size == n


# ----------------------------------------------------------- languages
def test_detect_pidgin():
    assert detect_language("Abeg wetin dey happen for our account?", ["en", "pcm"]) == "pcm"


def test_detect_pidgin():
    assert detect_language("abeg wetin dey happen with the sales", ["en", "pcm"]) == "pcm"


def test_detect_defaults_to_english():
    assert detect_language("Please summarise the quarterly report", ["en", "pcm"]) == "en"


def test_language_directive():
    d = language_directive("pcm")
    assert "Pidgin" in d
    assert language_directive("en") == ""


# ----------------------------------------------------------- memory
def test_memory_report_scoring():
    r = MemoryReport(peak_gb=2.1, budget_gb=7.0)
    assert r.within_budget
    assert r.efficiency_score == pytest.approx(70.0, abs=0.1)
    bad = MemoryReport(peak_gb=8.0, budget_gb=7.0)
    assert not bad.within_budget
    assert bad.efficiency_score == 0.0


def test_measure_peak_context():
    with measure_peak(budget_gb=7.0) as s:
        _ = [0] * 100000
    assert s.report.peak_gb >= 0


# ----------------------------------------------------------- agent (mock)
def _agent(tmp_path):
    cfg = load_config()
    cfg.rag.index_dir = str(tmp_path / "idx")
    cfg.rag.min_score = 0.0  # deterministic retrieval for the mock embedder
    embedder = MockEmbedder()
    idx = RagIndex(tmp_path / "idx")
    idx.build(ROOT / "data" / "corpus", embedder, cfg.rag.chunk_size, cfg.rag.chunk_overlap)
    cfg.embedding.path = str(tmp_path / "missing_embed.gguf")  # forces MockEmbedder
    cfg.rag.corpus_dir = str(ROOT / "data" / "corpus")
    return SabiAgent(cfg, MockChat(), idx)


def _collect(agent, message):
    """Return (full_text, events) from an agent run over the event stream."""
    events = list(agent.run(message))
    text = "".join(e["text"] for e in events if e.get("type") == "delta")
    return text, events


def test_agent_answers_with_context(tmp_path):
    agent = _agent(tmp_path)
    text, events = _collect(agent, "what is our refund policy?")
    assert len(text) > 0
    assert any(e["type"] == "sources" for e in events)  # RAG surfaced sources
    assert len(agent.history) == 2


def test_agent_uses_calc_tool(tmp_path):
    agent = _agent(tmp_path)
    text, events = _collect(agent, "what is 0.15 * 52000000 ?")
    tool_events = [e for e in events if e["type"] == "tool"]
    assert tool_events and tool_events[0]["name"] == "calc"
    assert "7800000" in tool_events[0]["summary"] or "7800000" in text


def test_identity_question_skips_documents(tmp_path):
    # Identity questions must NOT pull document sources (no doc can redefine Sabi)
    agent = _agent(tmp_path)
    _, events = _collect(agent, "who trained you and what model are you?")
    assert not any(e["type"] == "sources" for e in events)
    _, events2 = _collect(agent, "what is your name?")
    assert not any(e["type"] == "sources" for e in events2)


def test_identity_regex():
    from sabi.agent import _IDENTITY_RE
    for q in ["who are you", "what is your name", "who trained you?",
              "how were you trained", "what model are you", "are you mistral?",
              "who made you"]:
        assert _IDENTITY_RE.search(q), q
    assert not _IDENTITY_RE.search("what is our refund policy")
    assert not _IDENTITY_RE.search("summarise the sales report")


def test_explicit_language():
    from sabi.agent import _explicit_language
    assert _explicit_language("summarize in Pidgin please") == "pcm"
    assert _explicit_language("reply in English") == "en"
    assert _explicit_language("just answer normally") is None


def test_auto_aggregate_total_revenue():
    from sabi.tools import auto_aggregate
    res = auto_aggregate("what is our total revenue across all regions?", ROOT / "data" / "corpus")
    assert res and res["op"] == "sum"
    assert res["value_column"] == "revenue_ngn"
    assert res["result"] == pytest.approx(36690000.0)
    assert res["breakdown"]  # per-region rows present


def test_auto_aggregate_ignores_non_data():
    from sabi.tools import auto_aggregate
    assert auto_aggregate("what is our refund policy?", ROOT / "data" / "corpus") is None


def test_aggregate_xlsx(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    p = tmp_path / "sales.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["region", "amount"]); ws.append(["Lagos", 100]); ws.append(["Abuja", 250])
    wb.save(str(p))
    from sabi.tools import tool_aggregate
    res = tool_aggregate(file="sales.xlsx", column="amount", op="sum", corpus_dir=tmp_path)
    assert res["result"] == 350


def test_agent_saves_to_knowledge_base(tmp_path):
    cfg = load_config()
    corpus = tmp_path / "corpus"; corpus.mkdir()
    cfg.rag.corpus_dir = str(corpus)
    cfg.rag.index_dir = str(tmp_path / "idx")
    cfg.embedding.path = str(tmp_path / "missing.gguf")  # MockEmbedder
    idx = RagIndex(tmp_path / "idx")
    agent = SabiAgent(cfg, MockChat(), idx)
    events = list(agent.run("save this: our office closes at 5pm on Fridays"))
    assert any(e["type"] == "tool" and e["name"] == "save_to_knowledge_base" for e in events)
    kb = corpus / "knowledge_base.md"
    assert kb.exists() and "5pm" in kb.read_text()
    # it is now retrievable
    text, _ = _collect(agent, "when does the office close?")
    assert idx.size >= 1


def test_extract_expression():
    from sabi.tools import extract_expression
    assert extract_expression("what is 0.15 * 52000000?") == "0.15*52000000"
    assert extract_expression("(45000-38000)/38000") == "(45000-38000)/38000"
    assert extract_expression("hello how are you") is None
    assert extract_expression("revenue was 5000000") is None  # no operator


def test_embedder_ngram_quality():
    # char n-gram embedder should rank the on-topic chunk above an off-topic one
    emb = MockEmbedder()
    docs = emb.embed(["refund and returns policy for customers",
                      "working hours are 8:30 to 17:00"])
    q = emb.embed(["what is the refund policy"])[0]
    from sabi.embeddings import cosine_sim
    sims = cosine_sim(q, docs)
    assert sims[0] > sims[1]


# ----------------------------------------------------------- ingest + upload
def test_extract_text_plain(tmp_path):
    from sabi.ingest import extract_text
    p = tmp_path / "note.txt"
    p.write_text("Quarterly revenue grew by 12 percent.", encoding="utf-8")
    assert "Quarterly revenue" in extract_text(p)


def test_extract_text_docx(tmp_path):
    docx = pytest.importorskip("docx")  # python-docx
    p = tmp_path / "memo.docx"
    d = docx.Document()
    d.add_paragraph("Board meeting scheduled for Friday.")
    d.save(str(p))
    from sabi.ingest import extract_text
    assert "Board meeting" in extract_text(p)


def test_index_add_file_incremental(tmp_path):
    emb = MockEmbedder()
    idx = RagIndex(tmp_path / "idx")
    idx.build(ROOT / "data" / "corpus", emb, 400, 80)
    before = idx.size
    newdoc = tmp_path / "vendor_contract.txt"
    newdoc.write_text("The vendor warranty period is 24 months from delivery.", encoding="utf-8")
    added = idx.add_file(newdoc, emb, 400, 80)
    assert added >= 1
    assert idx.size == before + added
    # re-uploading the same file replaces, not duplicates
    again = idx.add_file(newdoc, emb, 400, 80)
    assert idx.size == before + again
    # the new content is retrievable
    qv = emb.embed(["what is the vendor warranty period"])[0]
    hits = idx.search(qv, top_k=3, min_score=0.0)
    assert any(c.source == "vendor_contract.txt" for c, _ in hits)


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))


# ----------------------------------------------------------- context budgeting
def test_budget_messages_trims_and_fits():
    from sabi.model import budget_messages
    cnt = lambda msgs: sum(len(m["content"]) for m in msgs)
    msgs = [
        {"role": "system", "content": "S" * 100},
        {"role": "user", "content": "U" * 500},
        {"role": "assistant", "content": "A" * 2000},
        {"role": "user", "content": "latest question"},
    ]
    trimmed, eff = budget_messages(msgs, n_ctx=1000, want_max=300, count_fn=cnt)
    assert trimmed[0]["role"] == "system"
    assert trimmed[-1]["content"] == "latest question"
    assert cnt(trimmed) + eff <= 1000
    assert eff >= 64


# ----------------------------------------------------------- excel export
def test_export_to_xlsx(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from sabi.tools import export_to_xlsx
    name = export_to_xlsx("Revenue: 100\nMargins improved this quarter", "Summary", tmp_path)
    assert name.endswith(".xlsx")
    wb = openpyxl.load_workbook(str(tmp_path / name))
    vals = [str(v) for row in wb.active.iter_rows(values_only=True) for v in row if v]
    assert any("Revenue" in v for v in vals)


def test_agent_creates_excel(tmp_path):
    pytest.importorskip("openpyxl")
    cfg = load_config()
    cfg.rag.enabled = False
    cfg.embedding.path = str(tmp_path / "x.gguf")
    agent = SabiAgent(cfg, MockChat(), None)
    agent.history += [
        {"role": "user", "content": "summarise the report"},
        {"role": "assistant", "content": "Summary: revenue grew, costs fell, and margins "
                                          "improved across all regions this quarter for the business."},
    ]
    events = list(agent.run("create an excel sheet of that summary"))
    tool = [e for e in events if e["type"] == "tool" and e["name"] == "create_excel"]
    assert tool, "expected a create_excel tool event"
    link = "".join(e["text"] for e in events if e["type"] == "delta")
    assert "/download/" in link and ".xlsx" in link
    # clean up the file the agent actually wrote
    from sabi.config import ROOT
    f = ROOT / "data" / "exports" / tool[0]["summary"]
    if f.exists():
        f.unlink()


# ----------------------------------------------------------- pivot / analysis
def test_pivot_table_2d():
    from sabi.tools import pivot_table
    piv = pivot_table("q1_sales_report.csv", index="region", value="revenue_ngn",
                      agg="sum", columns="month", corpus_dir=ROOT / "data" / "corpus")
    assert "error" not in piv
    assert piv["header"][0] == "region" and "Total" in piv["header"]
    assert len(piv["rows"]) == 4  # four regions


def test_auto_pivot_detection():
    from sabi.tools import auto_pivot
    piv = auto_pivot("make a pivot table of revenue by region and month", ROOT / "data" / "corpus")
    assert piv and piv["value"] == "revenue_ngn"
    assert piv["index"] == "region" and piv["columns"] == "month"


def test_analyze_table():
    from sabi.tools import analyze_table
    res = analyze_table("q1_sales_report.csv", ROOT / "data" / "corpus")
    cols = {s["column"] for s in res["stats"]}
    assert "revenue_ngn" in cols and "units_sold" in cols


def test_agent_stateless_history(tmp_path):
    cfg = load_config(); cfg.rag.enabled = False
    cfg.embedding.path = str(tmp_path / "x.gguf")
    agent = SabiAgent(cfg, MockChat(), None)
    h = [{"role": "user", "content": "hi"}, {"role": "assistant", "content": "hello there my friend"}]
    list(agent.run("tell me something useful", language="en", history=h))
    assert agent.history == []  # web mode: server keeps no per-conversation state


def test_agent_pivot_via_run(tmp_path):
    cfg = load_config()
    cfg.rag.index_dir = str(tmp_path / "idx")
    cfg.embedding.path = str(tmp_path / "x.gguf")
    cfg.rag.corpus_dir = str(ROOT / "data" / "corpus")
    idx = RagIndex(tmp_path / "idx")
    agent = SabiAgent(cfg, MockChat(), idx)
    events = list(agent.run("make a pivot table of revenue by region and month", history=[]))
    assert any(e["type"] == "tool" and e["name"] == "pivot_table" for e in events)
    text = "".join(e["text"] for e in events if e["type"] == "delta")
    assert "Total" in text and "|" in text  # rendered as a table


# ----------------------------------------------------------- deterministic table Q&A
def _corpus():
    return ROOT / "data" / "corpus"


def test_query_who_is_owing_lists_all():
    from sabi.tools import query_table
    q = query_table("who is currently owing?", _corpus())
    assert q and q["kind"] == "list"
    # both debtors present, paid customers absent
    assert "Mary Okafor" in q["markdown"] and "Blessing Eze" in q["markdown"]
    assert "John Ade" not in q["markdown"]
    assert len(q["rows"]) == 2


def test_query_count_owing():
    from sabi.tools import query_table
    q = query_table("how many people are owing?", _corpus())
    assert q and q["kind"] == "count" and q["count"] == 2


def test_query_highest_payment():
    from sabi.tools import query_table
    q = query_table("who made the highest payment?", _corpus())
    assert q and "Mary Okafor" in q["markdown"] and "400,000" in q["markdown"]


def test_query_total_owed():
    from sabi.tools import query_table
    q = query_table("what is the total amount owed altogether?", _corpus())
    assert q and "460,000" in q["markdown"]  # 320k + 140k


def test_pivot_not_triggered_without_keyword():
    from sabi.tools import auto_pivot
    assert auto_pivot("who is owing money by the way", _corpus()) is None


def test_extract_sheet_name():
    from sabi.tools import extract_sheet_name
    assert extract_sheet_name("create a new sheet and call it debtors") == "Debtors"
    assert extract_sheet_name("make a sheet for the debtors") == "Debtors"


def test_agent_creates_debtors_sheet(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    cfg = load_config()
    cfg.rag.index_dir = str(tmp_path / "idx")
    cfg.embedding.path = str(tmp_path / "x.gguf")
    cfg.rag.corpus_dir = str(_corpus())
    agent = SabiAgent(cfg, MockChat(), RagIndex(tmp_path / "idx"))
    events = list(agent.run("create a new sheet of debtors and call it Debtors with their balances", history=[]))
    tool = [e for e in events if e["type"] == "tool" and e["name"] == "create_excel"]
    assert tool
    text = "".join(e["text"] for e in events if e["type"] == "delta")
    assert "Debtors" in text and "/download/" in text
    name = ROOT / "data" / "exports" / text.split("/download/")[1].split(")")[0]
    wb = openpyxl.load_workbook(str(name))
    assert wb.active.title == "Debtors"
    vals = [str(v) for row in wb.active.iter_rows(values_only=True) for v in row if v]
    assert any("Mary Okafor" in v for v in vals) and any("Blessing Eze" in v for v in vals)
    name.unlink(missing_ok=True)


def test_languages_english_pidgin_only():
    from sabi.languages import detect_language, LANG_NAMES
    assert set(LANG_NAMES) == {"en", "pcm"}
    assert detect_language("abeg wetin dey happen", ["en", "pcm"]) == "pcm"
    assert detect_language("what is our total revenue", ["en", "pcm"]) == "en"


def test_compliance_report():
    from sabi.compliance import gather, render_markdown
    cfg = load_config()
    d = gather(cfg, None, 1.8)
    assert d["memory"]["budget_gb"] == 7.0
    assert any("offline" in c["label"].lower() for c in d["checks"])
    md = render_markdown(d)
    assert "Compliance" in md and "S_total" in md


# ----------------------------------------------------------- document awareness
def _doc_agent(tmp_path):
    cfg = load_config()
    cfg.rag.index_dir = str(tmp_path / "idx")
    cfg.embedding.path = str(tmp_path / "x.gguf")
    cfg.rag.corpus_dir = str(ROOT / "data" / "corpus")
    return SabiAgent(cfg, MockChat(), RagIndex(tmp_path / "idx"))


def test_doc_query_has_specific(tmp_path):
    a = _doc_agent(tmp_path)
    ans = a._doc_query("do you have access to the gworld sales document?")
    assert ans and ans.lower().startswith("yes") and "gworld" in ans.lower()


def test_doc_query_lists_all(tmp_path):
    a = _doc_agent(tmp_path)
    ans = a._doc_query("list all the documents you have access to")
    assert ans and "gworld_mobile_sales.xlsx" in ans.lower()


def test_doc_query_missing(tmp_path):
    a = _doc_agent(tmp_path)
    ans = a._doc_query("do you have access to the aura identity document?")
    assert ans and ans.lower().startswith("no")


def test_doc_query_ignores_content_questions(tmp_path):
    a = _doc_agent(tmp_path)
    # a content/data question must NOT be hijacked by the doc handler
    assert a._doc_query("summarise the gworld document") is None
    assert a._doc_query("who is owing in the sales file?") is None


def test_doc_query_via_run(tmp_path):
    a = _doc_agent(tmp_path)
    events = list(a.run("what documents do you have?", history=[]))
    text = "".join(e["text"] for e in events if e["type"] == "delta")
    assert "gworld_mobile_sales.xlsx" in text.lower()


def test_pidgin_detection_short_greeting():
    from sabi.languages import detect_language
    assert detect_language("how you dey?", ["en", "pcm"]) == "pcm"
    assert detect_language("how far boss", ["en", "pcm"]) == "pcm"
    assert detect_language("good morning, please summarise the report", ["en", "pcm"]) == "en"


# ----------------------------------------------------------- show table + filtering
def test_show_full_table_all_regions():
    from sabi.tools import show_table
    tv = show_table("show me the full q1 sales report table", ROOT / "data" / "corpus")
    assert tv and tv["rows"] == 12  # all 12 rows, not just Lagos
    for region in ("Lagos", "Abuja", "Port Harcourt", "Kano"):
        assert region in tv["markdown"]


def test_show_table_filtered_fuzzy_region():
    from sabi.tools import show_table
    # 'Porthacourt' (typo, no space) must still match 'Port Harcourt'
    tv = show_table("list all the sales from Porthacourt", ROOT / "data" / "corpus")
    assert tv and tv["rows"] == 3
    assert "Port Harcourt" in tv["markdown"] and "Lagos" not in tv["markdown"]


def test_single_region_total():
    from sabi.tools import auto_aggregate
    agg = auto_aggregate("what was the total revenue in Porthacourt?", ROOT / "data" / "corpus")
    assert agg and agg["filter_value"] == "Port Harcourt"
    assert agg["result"] == 7330000  # not the 36.69M grand total


def test_grand_total_unfiltered():
    from sabi.tools import auto_aggregate
    agg = auto_aggregate("what is our total revenue?", ROOT / "data" / "corpus")
    assert agg and not agg.get("filter_value")
    assert agg["result"] == 36690000


def test_count_owing_picks_right_table():
    from sabi.tools import query_table
    # must use the debtors spreadsheet, not q1_sales_report (which has no debtors)
    q = query_table("how many customers are owing?", ROOT / "data" / "corpus")
    assert q and q["count"] == 2 and "gworld" in q["file"].lower()


def test_show_table_ignores_summary_requests():
    from sabi.tools import show_table
    assert show_table("summarise the sales report", ROOT / "data" / "corpus") is None
